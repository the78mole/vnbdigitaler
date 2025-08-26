#!/usr/bin/env python3
"""BNetzA Rollout Excel to CSV Converter.

This module provides functionality to convert BNetzA rollout Excel reports
into clean CSV format, handling complex headers             # Progress logging every 10% or significant milestones
            progress = (i /            # Progress logging every 10% or significant milestones
            progress = (i / total_items) * PROGRESS_THRESHOLD
            show_progress = (
                i == 1
                or i == total_items  # First item (0%)
                or (  # Last item (100%)
                    total_items >= MIN_ROW_COUNT
                    and progress >= MIN_ROW_COUNT
                    and (progress % PROGRESS_INTERVAL) <= (PROGRESS_THRESHOLD / total_items)
                )  # Every 10%
            )) * PROGRESS_THRESHOLD
            show_progress = (
                i == 1
                or i == total_rows  # First item (0%)
                or (  # Last item (100%)
                    total_rows >= MIN_ROW_COUNT
                    and progress >= MIN_ROW_COUNT
                    and (progress % PROGRESS_INTERVAL) <= (PROGRESS_THRESHOLD / total_rows)
                )  # Every 10%
            )ata formats.

The Excel files typically have:
- Row 1-2: Human-readable headers (ignored)
- Row 3: Column headers starting from row 3
- Column A: Company names
- Column C: Quota values (either numeric 0.00-1.00 or strings like "0% (Stichtag 31.12.2024)")

Output CSV format:
- Company,Ausstattungsquote,Stichtag
- Where Ausstattungsquote is always a float 0.00-1.00
- And Stichtag is the reference date (either from quarter or extracted from string)
"""

import csv
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Constants for magic value compliance
PROGRESS_INTERVAL = 10.0  # Progress logging interval in percentage
PROGRESS_THRESHOLD = 100.0  # Total progress percentage
PERCENTAGE_DIVISOR = 100.0  # Divisor for percentage conversion
MIN_ROW_COUNT = 10  # Minimum row count for progress logging
MIN_QUOTA_VALUE = 0.0  # Minimum valid quota value
MAX_QUOTA_VALUE = 1.0  # Maximum valid quota value# Setup module logger
logger = logging.getLogger(__name__)


class BNetzARolloutXlsx2CsvConverter:
    """Converter for BNetzA rollout Excel reports to clean CSV format.

    This class handles the complex structure of BNetzA Excel reports:
    - Skips human-readable headers in rows 1-2
    - Extracts data starting from row 3
    - Processes mixed data formats in quota column
    - Standardizes output to consistent CSV format
    """

    def __init__(self) -> None:
        """Initialize the converter."""
        self.logger = logging.getLogger("xlsx_converter")

        # Regex patterns for parsing quota strings
        self.quota_pattern = re.compile(
            r"(\d+(?:[,\.]\d+)?)\s*%\s*\((?:Stichtag\s+)?(\d{1,2})\.(\d{1,2})\.(\d{4})\)",
            re.IGNORECASE,
        )

        # Standard quarter end dates
        self.quarter_end_dates = {
            1: (31, 3),  # Q1: March 31
            2: (30, 6),  # Q2: June 30
            3: (30, 9),  # Q3: September 30
            4: (31, 12),  # Q4: December 31
        }

    def convert_xlsx_to_csv(
        self,
        excel_path: str | Path,
        csv_path: str | Path,
        quarter: int | None = None,
        year: int | None = None,
    ) -> dict[str, Any]:
        """Convert BNetzA Excel report to clean CSV format.

        Args:
            excel_path: Path to the input Excel file
            csv_path: Path for the output CSV file
            quarter: Quarter number (1-4) for default reference date
            year: Year for default reference date

        Returns:
            Dictionary with conversion statistics and metadata

        Raises:
            FileNotFoundError: If Excel file doesn't exist
            ValueError: If Excel file format is invalid
        """
        excel_path = Path(excel_path)
        csv_path = Path(csv_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        self.logger.info(f"🔄 Converting Excel to CSV: {excel_path.name}")
        self.logger.info(f"📁 Output CSV: {csv_path}")

        try:
            # Load Excel workbook
            workbook = load_workbook(excel_path, read_only=True, data_only=True)

            # Get the first worksheet (should contain the rollout data)
            if not workbook.worksheets:
                raise ValueError("Excel file contains no worksheets")

            worksheet = workbook.worksheets[0]
            self.logger.info(f"📊 Processing worksheet: {worksheet.title}")

            # Extract data starting from row 3 (skip headers)
            raw_data = self._extract_raw_data(worksheet)
            self.logger.info(f"📋 Extracted {len(raw_data)} raw data rows")

            # Determine default reference date from quarter/year
            default_reference_date = self._get_default_reference_date(quarter, year)

            # Process and clean the data
            processed_data = self._process_data(raw_data, default_reference_date)
            self.logger.info(f"✅ Processed {len(processed_data)} valid data rows")

            # Write to CSV
            self._write_csv(processed_data, csv_path)

            # Generate statistics
            stats = self._generate_statistics(
                raw_data, processed_data, default_reference_date
            )

            self.logger.info("🎉 Conversion completed successfully")
            self.logger.info(f"📊 Statistics: {stats}")

            return stats

        except Exception as e:
            self.logger.error(f"❌ Conversion failed: {e}")
            raise
        finally:
            if "workbook" in locals():
                workbook.close()

    def _extract_raw_data(self, worksheet: Any) -> list[dict[str, Any]]:
        """Extract raw data from Excel worksheet starting from row 3.

        Args:
            worksheet: OpenPyXL worksheet object

        Returns:
            List of dictionaries with raw data from columns A and C
        """
        raw_data = []
        total_rows = worksheet.max_row - 2  # Subtract 2 for header rows

        self.logger.info(f"📊 Extracting data from {total_rows} rows...")

        # Start from row 3 (1-indexed), skip header rows
        for i, row_num in enumerate(range(3, worksheet.max_row + 1), 1):
            company_cell = worksheet[f"A{row_num}"]
            quota_cell = worksheet[f"C{row_num}"]

            # Get cell values
            company_name = company_cell.value
            quota_value = quota_cell.value

            # Skip empty rows
            if not company_name and not quota_value:
                continue

            # Skip rows with only company name but no quota
            if company_name and not quota_value:
                self.logger.debug(
                    f"Skipping row {row_num}: No quota value for '{company_name}'"
                )
                continue

            # Skip rows with only quota but no company name
            if quota_value and not company_name:
                self.logger.debug(
                    f"Skipping row {row_num}: No company name for quota '{quota_value}'"
                )
                continue

            raw_data.append(
                {
                    "row_number": row_num,
                    "company_name": str(company_name).strip() if company_name else "",
                    "quota_raw": quota_value,
                    "quota_original": str(quota_value)
                    if quota_value is not None
                    else "",
                }
            )

            # Progress logging every 10% or significant milestones
            progress = (i / total_rows) * 100
            show_progress = i in (1, total_rows) or (
                total_rows >= MIN_ROW_COUNT
                and progress >= MIN_ROW_COUNT
                and progress % PROGRESS_INTERVAL <= PROGRESS_THRESHOLD / total_rows
            )  # Every 10%
            if show_progress:
                self.logger.info(
                    f"📊 Data extraction progress: {progress:.0f}% ({i}/{total_rows})"
                )

        return raw_data

    def _get_default_reference_date(
        self, quarter: int | None, year: int | None
    ) -> datetime | None:
        """Get the default reference date based on quarter and year.

        Args:
            quarter: Quarter number (1-4)
            year: Year

        Returns:
            Default reference date or None if quarter/year not provided
        """
        if not quarter or not year:
            return None

        if quarter not in self.quarter_end_dates:
            self.logger.warning(f"Invalid quarter {quarter}, using Q1 as fallback")
            quarter = 1

        day, month = self.quarter_end_dates[quarter]
        return datetime(year, month, day)

    def _process_data(
        self, raw_data: list[dict[str, Any]], default_reference_date: datetime | None
    ) -> list[dict[str, Any]]:
        """Process raw data into clean format.

        Args:
            raw_data: List of raw data dictionaries
            default_reference_date: Default reference date for the quarter

        Returns:
            List of processed data dictionaries
        """
        processed_data = []
        total_items = len(raw_data)

        self.logger.info(f"🔄 Processing {total_items} data items...")

        for i, item in enumerate(raw_data, 1):
            try:
                result = self._process_quota_value(
                    item["quota_raw"], default_reference_date
                )

                if result:
                    processed_data.append(
                        {
                            "company_name": item["company_name"],
                            "ausstattungsquote": result["quota"],
                            "stichtag": result["reference_date"].strftime("%Y-%m-%d"),
                            "row_number": item["row_number"],
                            "original_value": item["quota_original"],
                            "parsing_method": result["method"],
                        }
                    )
                else:
                    self.logger.warning(
                        f"Could not process quota for {item['company_name']}: '{item['quota_original']}'"
                    )

            except Exception as e:
                self.logger.error(
                    f"Error processing row {item['row_number']} ({item['company_name']}): {e}"
                )

            # Progress logging every 10% or significant milestones
            progress = (i / total_items) * 100
            show_progress = i in (1, total_items) or (
                total_items >= MIN_ROW_COUNT
                and progress >= MIN_ROW_COUNT
                and progress % PROGRESS_INTERVAL <= PROGRESS_THRESHOLD / total_items
            )  # Every 10%
            if show_progress:
                self.logger.info(
                    f"🔄 Data processing progress: {progress:.0f}% ({i}/{total_items})"
                )

        return processed_data

    def _process_quota_value(  # noqa: PLR0911
        self, quota_value: Any, default_reference_date: datetime | None
    ) -> dict[str, Any] | None:
        """Process a single quota value into standardized format.

        Args:
            quota_value: Raw quota value from Excel cell
            default_reference_date: Default reference date if not specified in value

        Returns:
            Dictionary with quota and reference_date, or None if invalid
        """
        if quota_value is None:
            return None

        # Case 1: Numeric value (0.00 - 1.00)
        if isinstance(quota_value, int | float):
            if MIN_QUOTA_VALUE <= quota_value <= MAX_QUOTA_VALUE:
                if not default_reference_date:
                    self.logger.warning(
                        "Numeric quota found but no default reference date provided"
                    )
                    return None
                return {
                    "quota": float(quota_value),
                    "reference_date": default_reference_date,
                    "method": "numeric",
                }
            else:
                self.logger.warning(f"Numeric quota out of range: {quota_value}")
                return None

        # Case 2: String value (e.g., "0% (Stichtag 31.12.2024)")
        if isinstance(quota_value, str):
            quota_str = quota_value.strip()

            # Try to parse string with regex
            match = self.quota_pattern.match(quota_str)
            if match:
                percentage_str, day_str, month_str, year_str = match.groups()

                try:
                    # Convert percentage to decimal (handle German comma notation)
                    percentage_str = percentage_str.replace(",", ".")
                    percentage = float(percentage_str)
                    quota = percentage / PERCENTAGE_DIVISOR

                    # Parse date components
                    day = int(day_str)
                    month = int(month_str)
                    year = int(year_str)

                    # Validate quota range
                    if not (MIN_QUOTA_VALUE <= quota <= MAX_QUOTA_VALUE):
                        self.logger.warning(
                            f"Quota out of range in string: {quota_str}"
                        )
                        return None

                    # Create reference date
                    reference_date = datetime(year, month, day)

                    return {
                        "quota": quota,
                        "reference_date": reference_date,
                        "method": "string_parsed",
                    }

                except (ValueError, OverflowError) as e:
                    self.logger.warning(
                        f"Could not parse date/quota from string '{quota_str}': {e}"
                    )
                    return None

            # Try to parse as simple percentage without date
            simple_percentage_match = re.match(r"(\d+(?:[,\.]\d+)?)\s*%", quota_str)
            if simple_percentage_match:
                try:
                    percentage_str = simple_percentage_match.group(1).replace(",", ".")
                    percentage = float(percentage_str)
                    quota = percentage / PERCENTAGE_DIVISOR

                    if (
                        MIN_QUOTA_VALUE <= quota <= MAX_QUOTA_VALUE
                        and default_reference_date
                    ):
                        return {
                            "quota": quota,
                            "reference_date": default_reference_date,
                            "method": "percentage_only",
                        }
                except ValueError:
                    pass

            self.logger.warning(f"Could not parse quota string: '{quota_str}'")
            return None

        # Case 3: Unknown type
        self.logger.warning(
            f"Unknown quota value type: {type(quota_value)} - {quota_value}"
        )
        return None

    def _write_csv(self, processed_data: list[dict[str, Any]], csv_path: Path) -> None:
        """Write processed data to CSV file.

        Args:
            processed_data: List of processed data dictionaries
            csv_path: Output CSV file path
        """
        csv_path.parent.mkdir(parents=True, exist_ok=True)

        with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
            fieldnames = ["company_name", "ausstattungsquote", "stichtag"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            # Write header
            writer.writeheader()

            # Write data rows
            for item in processed_data:
                writer.writerow(
                    {
                        "company_name": item["company_name"],
                        "ausstattungsquote": item["ausstattungsquote"],
                        "stichtag": item["stichtag"],
                    }
                )

        self.logger.info(f"📁 CSV file written: {csv_path}")

    def _generate_statistics(
        self,
        raw_data: list[dict[str, Any]],
        processed_data: list[dict[str, Any]],
        default_reference_date: datetime | None,
    ) -> dict[str, Any]:
        """Generate conversion statistics.

        Args:
            raw_data: Original raw data
            processed_data: Successfully processed data
            default_reference_date: Default reference date used

        Returns:
            Statistics dictionary
        """
        stats: dict[str, Any] = {
            "input_file_rows": len(raw_data),
            "output_csv_rows": len(processed_data),
            "rows_skipped": len(raw_data) - len(processed_data),
            "default_reference_date": default_reference_date.strftime("%Y-%m-%d")
            if default_reference_date
            else None,
            "parsing_methods": {},
            "reference_dates_used": {},
            "quota_range": {"min": None, "max": None, "avg": None},
        }

        if processed_data:
            # Count parsing methods
            for item in processed_data:
                method = str(item.get("parsing_method", "unknown"))
                if "parsing_methods" not in stats:
                    stats["parsing_methods"] = {}
                stats["parsing_methods"][method] = (
                    stats["parsing_methods"].get(method, 0) + 1
                )

            # Count reference dates
            for item in processed_data:
                ref_date = str(item["stichtag"])
                if "reference_dates_used" not in stats:
                    stats["reference_dates_used"] = {}
                stats["reference_dates_used"][ref_date] = (
                    stats["reference_dates_used"].get(ref_date, 0) + 1
                )

            # Calculate quota statistics
            quotas = [float(item["ausstattungsquote"]) for item in processed_data]
            if quotas:
                if "quota_range" not in stats:
                    stats["quota_range"] = {}
                stats["quota_range"]["min"] = min(quotas)
                stats["quota_range"]["max"] = max(quotas)
                stats["quota_range"]["avg"] = sum(quotas) / len(quotas)

        return stats

    def convert_multiple_files(
        self,
        input_dir: str | Path,
        output_dir: str | Path,
        file_pattern: str = "*.xlsx",
    ) -> dict[str, dict[str, Any]]:
        """Convert multiple Excel files to CSV format.

        Args:
            input_dir: Directory containing Excel files
            output_dir: Directory for output CSV files
            file_pattern: Glob pattern for Excel files

        Returns:
            Dictionary mapping filenames to their conversion statistics
        """
        input_dir = Path(input_dir)
        output_dir = Path(output_dir)

        output_dir.mkdir(parents=True, exist_ok=True)

        results = {}
        excel_files = list(input_dir.glob(file_pattern))

        self.logger.info(
            f"🔄 Converting {len(excel_files)} Excel files from {input_dir}"
        )

        for excel_file in excel_files:
            try:
                # Generate CSV filename
                csv_file = output_dir / f"{excel_file.stem}.csv"

                # Extract quarter/year from filename if possible
                quarter, year = self._extract_quarter_year_from_filename(
                    excel_file.name
                )

                # Convert file
                stats = self.convert_xlsx_to_csv(excel_file, csv_file, quarter, year)
                stats["source_file"] = str(excel_file)
                stats["output_file"] = str(csv_file)

                results[excel_file.name] = stats

            except Exception as e:
                self.logger.error(f"❌ Failed to convert {excel_file.name}: {e}")
                results[excel_file.name] = {"error": str(e)}

        self.logger.info(
            f"🎉 Batch conversion completed: {len(results)} files processed"
        )
        return results

    def _extract_quarter_year_from_filename(
        self, filename: str
    ) -> tuple[int | None, int | None]:
        """Extract quarter and year from filename.

        Args:
            filename: Excel filename

        Returns:
            Tuple of (quarter, year) or (None, None) if not found
        """
        # Pattern for "Roll-out-Quoten_Q1_2025.xlsx"
        match = re.search(r"Q(\d)_(\d{4})", filename, re.IGNORECASE)
        if match:
            quarter = int(match.group(1))
            year = int(match.group(2))
            return quarter, year

        # Pattern for "2025-Q1" or similar
        match = re.search(r"(\d{4})[_-]Q(\d)", filename, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            quarter = int(match.group(2))
            return quarter, year

        return None, None
